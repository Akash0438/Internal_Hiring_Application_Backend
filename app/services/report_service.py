"""
report_service.py
─────────────────
Generates a daily Excel report of all candidates (with their assignments and
feedback) and emails it to every active Portfolio Manager.

Schedule: 5:00 PM IST = 11:30 AM UTC  (IST is UTC+5:30)

The report contains three sheets:
  1. Summary       — one row per candidate with current status
  2. Assignments   — one row per assignment
  3. Feedback      — one row per feedback entry
"""
from __future__ import annotations

import io
import logging
import smtplib
from datetime import datetime, timezone
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config import settings

logger = logging.getLogger(__name__)

# ── Colour palette ─────────────────────────────────────────────────────────────
_HEADER_FILL = PatternFill("solid", fgColor="1D4ED8")   # blue
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL    = PatternFill("solid", fgColor="EFF6FF")   # light-blue row stripe
_BORDER_SIDE = Side(style="thin", color="BFDBFE")
_CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE,
    top=_BORDER_SIDE, bottom=_BORDER_SIDE,
)

STATUS_COLORS = {
    "NEW":                "E5E7EB",
    "ASSIGNED":           "BFDBFE",
    "INTERVIEW_SCHEDULED":"DDD6FE",
    "INTERVIEW_COMPLETED":"C7D2FE",
    "FEEDBACK_SUBMITTED": "FDE68A",
    "UNDER_REVIEW":       "FED7AA",
    "APPROVED":           "BBF7D0",
    "REJECTED":           "FECACA",
    "ON_HOLD":            "FEF3C7",
}


# ── Helper ──────────────────────────────────────────────────────────────────────

def _write_header(ws, columns: list[str]) -> None:
    ws.append(columns)
    for col_idx, _ in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = _HEADER_FILL
        cell.font   = _HEADER_FONT
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22


def _style_row(ws, row_idx: int, num_cols: int, alt: bool, status: str | None = None) -> None:
    fill = PatternFill("solid", fgColor=STATUS_COLORS.get(status or "", "FFFFFF")) if status else (
        _ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    )
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill   = fill
        cell.border = _CELL_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=True)


def _autofit(ws, min_width: int = 12, max_width: int = 40) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_width, min(max_len + 2, max_width))


# ── Main report builder ────────────────────────────────────────────────────────

async def build_report() -> bytes:
    """
    Query MongoDB and return a fully styled Excel workbook as raw bytes.
    Import models inside the function to avoid circular imports.
    """
    from app.models.user import User, Role
    from app.models.candidate import Candidate
    from app.models.interview_assignment import InterviewAssignment
    from app.models.interview_feedback import InterviewFeedback

    wb = openpyxl.Workbook()
    generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # ── Lookup maps ───────────────────────────────────────────────────────────
    all_users = await User.find_all().to_list()
    user_map: dict[str, str] = {str(u.id): u.name for u in all_users}

    all_candidates = await Candidate.find_all().sort("created_at").to_list()
    cand_map: dict[str, Candidate] = {str(c.id): c for c in all_candidates}

    all_assignments = await InterviewAssignment.find_all().sort("assigned_date").to_list()
    assign_map: dict[str, list[InterviewAssignment]] = {}
    for a in all_assignments:
        assign_map.setdefault(str(a.candidate_id), []).append(a)

    all_feedback = await InterviewFeedback.find_all().to_list()
    fb_map: dict[str, InterviewFeedback] = {str(fb.assignment_id): fb for fb in all_feedback}

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 1 — Summary
    # ──────────────────────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.freeze_panes = "A2"

    cols1 = [
        "Emp ID", "Candidate Name", "Email", "Phone",
        "Position", "Pay Band", "Location", "Experience",
        "Status", "Created By (PM)", "Interviewer", "Feedback Rating",
        "Recommendation", "Created At", "Last Updated",
    ]
    _write_header(ws1, cols1)

    for i, c in enumerate(all_candidates, 2):
        # Find latest assignment for this candidate
        assigns = assign_map.get(str(c.id), [])
        latest_assign = assigns[-1] if assigns else None
        hm_name = user_map.get(str(latest_assign.hiring_manager_id), "—") if latest_assign else "—"
        fb = fb_map.get(str(latest_assign.id)) if latest_assign else None

        row = [
            c.employee_id,
            c.candidate_name,
            c.email,
            c.phone or "—",
            c.position,
            c.pay_band_level or "—",
            c.current_location or "—",
            c.experience or "—",
            c.status.value,
            user_map.get(str(c.created_by_id), "—"),
            hm_name,
            str(fb.rating) + "/5" if fb else "—",
            fb.recommendation.value.replace("_", " ").title() if fb else "—",
            c.created_at.strftime("%Y-%m-%d") if c.created_at else "—",
            c.updated_at.strftime("%Y-%m-%d %H:%M") if c.updated_at else "—",
        ]
        ws1.append(row)
        _style_row(ws1, i, len(cols1), alt=False, status=c.status.value)

    _autofit(ws1)

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 2 — Assignments
    # ──────────────────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Assignments")
    ws2.freeze_panes = "A2"

    cols2 = [
        "Emp ID", "Candidate Name", "Position", "Pay Band",
        "Portfolio Manager", "Interviewer",
        "Assigned Date", "Assignment Status", "Self-Assigned",
    ]
    _write_header(ws2, cols2)

    for i, a in enumerate(all_assignments, 2):
        c = cand_map.get(str(a.candidate_id))
        row = [
            c.employee_id if c else "—",
            c.candidate_name if c else "—",
            c.position if c else "—",
            c.pay_band_level or "—" if c else "—",
            user_map.get(str(a.portfolio_manager_id), "—"),
            user_map.get(str(a.hiring_manager_id), "—"),
            a.assigned_date.strftime("%Y-%m-%d") if a.assigned_date else "—",
            a.status.value,
            "Yes" if a.is_self_assigned else "No",
        ]
        ws2.append(row)
        _style_row(ws2, i, len(cols2), alt=i % 2 == 0)

    _autofit(ws2)

    # ──────────────────────────────────────────────────────────────────────────
    # Sheet 3 — Feedback
    # ──────────────────────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Feedback")
    ws3.freeze_panes = "A2"

    cols3 = [
        "Emp ID", "Candidate Name", "Position", "Pay Band",
        "Interviewer", "Rating", "Recommendation", "Feedback", "Submitted At",
    ]
    _write_header(ws3, cols3)

    for i, fb in enumerate(all_feedback, 2):
        a = next((x for x in all_assignments if x.id == fb.assignment_id), None)
        c = cand_map.get(str(a.candidate_id)) if a else None
        row = [
            c.employee_id if c else "—",
            c.candidate_name if c else "—",
            c.position if c else "—",
            c.pay_band_level or "—" if c else "—",
            user_map.get(str(a.hiring_manager_id), "—") if a else "—",
            f"{fb.rating}/5",
            fb.recommendation.value.replace("_", " ").title(),
            fb.feedback,
            fb.submitted_at.strftime("%Y-%m-%d %H:%M") if fb.submitted_at else "—",
        ]
        ws3.append(row)
        _style_row(ws3, i, len(cols3), alt=i % 2 == 0)

    _autofit(ws3)

    # ── Metadata sheet ────────────────────────────────────────────────────────
    ws_meta = wb.create_sheet("Report Info")
    ws_meta.append(["Interview Management Platform — Daily Report"])
    ws_meta.append(["Generated at (UTC)", generated_at])
    ws_meta.append(["Total Candidates", len(all_candidates)])
    ws_meta.append(["Total Assignments", len(all_assignments)])
    ws_meta.append(["Total Feedback Entries", len(all_feedback)])
    ws_meta["A1"].font = Font(bold=True, size=13, color="1D4ED8")
    ws_meta.column_dimensions["A"].width = 35
    ws_meta.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Email sender ───────────────────────────────────────────────────────────────

def send_report_email(to: str, name: str, report_bytes: bytes, date_str: str) -> None:
    """Send the Excel report to a single Portfolio Manager."""
    if not settings.GMAIL_USER or not settings.GMAIL_APP_PASSWORD:
        logger.warning("Gmail credentials not set — skipping daily report email to %s", to)
        return

    msg = MIMEMultipart("mixed")
    msg["Subject"] = f"Daily Interview Report — {date_str}"
    msg["From"]    = settings.FROM_EMAIL or settings.GMAIL_USER
    msg["To"]      = to

    body = MIMEText(f"""
    <div style="font-family:sans-serif;max-width:540px;margin:auto">
      <div style="background:#1d4ed8;padding:20px 24px;border-radius:8px 8px 0 0">
        <h2 style="margin:0;color:#fff;font-size:18px">Daily Interview Report</h2>
      </div>
      <div style="padding:24px;border:1px solid #e5e7eb;border-top:0;border-radius:0 0 8px 8px">
        <p style="margin-top:0">Hi <strong>{name}</strong>,</p>
        <p>Please find attached the end-of-day interview report for <strong>{date_str}</strong>.</p>
        <p>The report includes:</p>
        <ul style="font-size:14px;line-height:1.8">
          <li><strong>Summary</strong> — All candidates with current status, interviewer and feedback rating</li>
          <li><strong>Assignments</strong> — All interview assignments with PM and interviewer details</li>
          <li><strong>Feedback</strong> — All submitted feedback entries</li>
        </ul>
        <p style="font-size:12px;color:#6b7280;margin-top:24px">
          This is an automated end-of-day report sent at 5:00 PM IST.<br>
          Interview Management Platform
        </p>
      </div>
    </div>
    """, "html")
    msg.attach(body)

    attachment = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    attachment.set_payload(report_bytes)
    encoders.encode_base64(attachment)
    attachment.add_header(
        "Content-Disposition",
        f'attachment; filename="interview_report_{date_str}.xlsx"',
    )
    msg.attach(attachment)

    try:
        with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as smtp:
            smtp.ehlo()
            smtp.starttls()
            smtp.login(settings.GMAIL_USER, settings.GMAIL_APP_PASSWORD.replace(" ", ""))
            smtp.sendmail(msg["From"], [to], msg.as_string())
        logger.info("Daily report sent to %s (%s)", name, to)
    except Exception as exc:
        logger.error("Failed to send daily report to %s: %s", to, exc)


# ── Scheduled job ──────────────────────────────────────────────────────────────

async def send_daily_report() -> None:
    """
    Entry point called by APScheduler at 11:30 UTC (5:00 PM IST).
    Builds the report once and emails it to every active Portfolio Manager.
    """
    from app.models.user import User, Role

    date_str = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")
    logger.info("=== Daily report job started for %s ===", date_str)

    try:
        portfolio_managers = await User.find(
            User.role == Role.MAIN_MANAGER,
            User.is_active == True,  # noqa: E712
        ).to_list()

        if not portfolio_managers:
            logger.warning("No active Portfolio Managers found — skipping daily report.")
            return

        logger.info("Building Excel report...")
        report_bytes = await build_report()
        logger.info("Report built: %d bytes", len(report_bytes))

        for pm in portfolio_managers:
            send_report_email(pm.email, pm.name, report_bytes, date_str)

        logger.info("=== Daily report sent to %d PM(s) ===", len(portfolio_managers))

    except Exception as exc:
        logger.error("Daily report job failed: %s", exc, exc_info=True)
